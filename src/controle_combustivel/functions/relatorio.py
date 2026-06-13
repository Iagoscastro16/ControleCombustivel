from database.connection import get_connection
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

MESES_ABREV = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]


def gerar_relatorio(ano, mostrar_inativo=False):
    try:
        if mostrar_inativo == 1:
            with get_connection() as conn:
                cursor = conn.execute('''
                    SELECT v.nome, v.categoria, strftime('%m', a.data) as mes, sum(a.valor)
                    FROM abastecimentos a
                    JOIN veiculos v ON a.veiculo_id = v.id
                    WHERE strftime('%Y', a.data) = ?
                    GROUP BY v.nome, mes
                    ORDER BY categoria, v.nome, mes
                ''', (ano,))
                return cursor.fetchall()
        else:
            with get_connection() as conn:
                cursor = conn.execute('''
                    SELECT v.nome, v.categoria, strftime('%m', a.data) as mes, sum(a.valor)
                    FROM abastecimentos a
                    JOIN veiculos v ON a.veiculo_id = v.id
                    WHERE strftime('%Y', a.data) = ?
                    AND v.ativo = 1
                    GROUP BY v.nome, mes
                    ORDER BY categoria, v.nome, mes
                ''', (ano,))
                return cursor.fetchall()
    except Exception as e:
        return {"success": False, "message": f"Erro ao gerar relatório: {e}"}


def exportar_excel(dados, ano, caminho):
    wb = Workbook()
    ws = wb.active
    ws.title = f"combustivel {ano}"

    veiculos    = dados["veiculos"]
    totais_mes  = dados["totais_mes"]

    # ── Cores ─────────────────────────────────────────────────
    COR_HEADER  = "374151"   # cinza médio — cabeçalho, legível em P&B
    COR_TOTAL   = "FED7AA"   # laranja claro — total mês
    COR_CAT     = "E5E7EB"   # cinza claro  — separador categoria
    COR_AH_G    = "F3F4F6"   # quase branco — A/H geral
    COR_BRANCO  = "FFFFFF"
    COR_PRETO   = "111827"   # texto escuro para fundos claros
    COR_LINHA_A = "FFFFFF"
    COR_LINHA_B = "F9FAFB"
    COR_SUBIR   = "DC2626"   # vermelho forte — mais visível em P&B
    COR_CAIR    = "16A34A"   # verde forte   — mais visível em P&B

    # A/V% calculado UMA vez com total anual completo (12 meses)
    total_geral_anual = sum(sum(v for v in valores if v) for _, _, valores in veiculos)

    # ── Helpers de estilo ─────────────────────────────────────
    def _fill(cell, cor):
        cell.fill = PatternFill(fill_type="solid", fgColor=cor)

    def _header(cell, valor, align="center"):
        cell.value = valor
        cell.font  = Font(bold=True, color=COR_BRANCO, size=10)
        _fill(cell, COR_HEADER)
        cell.alignment = Alignment(horizontal=align, vertical="center")

    def _total(cell, valor=None, fmt=None, bold=True, align="center"):
        cell.value = valor
        cell.font  = Font(bold=bold, color=COR_PRETO, size=10)
        _fill(cell, COR_TOTAL)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            cell.number_format = fmt

    def _ah_style(cell, valor=None, fmt=None, cor_fundo=None, cor_texto=None, bold=False):
        cell.value = valor
        cor_txt = cor_texto if cor_texto else COR_PRETO
        cell.font  = Font(bold=bold, color=cor_txt, size=9)
        _fill(cell, cor_fundo or COR_AH_G)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            cell.number_format = fmt

    def _calcular_ah(atual, anterior):
        if not anterior or anterior == 0:
            return None
        return (atual / anterior - 1) if atual else -1.0

    # ── Escreve um bloco semestral ────────────────────────────
    def escrever_bloco(linha_inicio, meses_idx, meses_nomes):
        linha  = linha_inicio
        n_cols = len(meses_idx)      # 6
        col_total = n_cols + 2       # coluna TOTAL ANO (1-based: nome=1, meses=2..7, total=8)
        col_av    = n_cols + 3       # coluna A/V% (col 9)

        # Título
        ws.merge_cells(f"A{linha}:I{linha}")
        ws[f"A{linha}"].value     = f"COMBUSTIVEL {ano} (POSTO/CAIXA)"
        ws[f"A{linha}"].font      = Font(bold=True, size=13)
        ws[f"A{linha}"].alignment = Alignment(horizontal="center")
        linha += 1

        # Cabeçalho
        _header(ws.cell(row=linha, column=1), "VEÍCULOS/PERÍODO", align="left")
        for col, mes in enumerate(meses_nomes, start=2):
            _header(ws.cell(row=linha, column=col), mes)
        _header(ws.cell(row=linha, column=col_total), "TOTAL ANO")
        _header(ws.cell(row=linha, column=col_av), "A/V %")
        linha += 1

        # Veículos
        categoria_atual = None
        for i_veiculo, (nome, categoria, valores) in enumerate(veiculos):
            total_veiculo_anual = sum(v for v in valores if v)

            # Separador de categoria
            if categoria != categoria_atual:
                categoria_atual = categoria
                cell = ws.cell(row=linha, column=1, value=categoria.upper())
                cell.font      = Font(bold=True, color=COR_PRETO, size=9)
                cell.fill      = PatternFill(fill_type="solid", fgColor=COR_CAT)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                ws.merge_cells(f"A{linha}:I{linha}")
                ws.row_dimensions[linha].height = 16
                linha += 1

            cor_bg = COR_LINHA_A if i_veiculo % 2 == 0 else COR_LINHA_B

            # ── Linha de valores ──────────────────────────────
            cell_nome            = ws.cell(row=linha, column=1, value=nome)
            cell_nome.font       = Font(size=9)
            cell_nome.fill       = PatternFill(fill_type="solid", fgColor=cor_bg)
            cell_nome.alignment  = Alignment(horizontal="left", vertical="center")

            total_veiculo = 0
            for col, idx in enumerate(meses_idx, start=2):
                val  = valores[idx] if valores[idx] else None
                cell = ws.cell(row=linha, column=col)
                cell.value         = val
                cell.number_format = 'R$ #,##0.00'
                cell.alignment     = Alignment(horizontal="center", vertical="center")
                _fill(cell, cor_bg)
                if val:
                    total_veiculo += val

            # A/V% anual
            av = (total_veiculo_anual / total_geral_anual) if total_geral_anual and total_veiculo else None

            # TOTAL ANO
            cell_total               = ws.cell(row=linha, column=col_total)
            cell_total.value         = total_veiculo_anual if total_veiculo_anual else None
            cell_total.number_format = 'R$ #,##0.00'
            cell_total.font          = Font(bold=True, size=9)
            cell_total.alignment     = Alignment(horizontal="center", vertical="center")
            _fill(cell_total, cor_bg)

            cell_av_v               = ws.cell(row=linha, column=col_av)
            cell_av_v.value         = av
            cell_av_v.number_format = '0.0%'
            cell_av_v.alignment     = Alignment(horizontal="center", vertical="center")
            _fill(cell_av_v, cor_bg)

            ws.row_dimensions[linha].height = 15
            linha += 1

            # ── Linha A/H% por veículo ────────────────────────
            ws.cell(row=linha, column=1).value = ""
            _fill(ws.cell(row=linha, column=1), COR_AH_G)

            for col, idx in enumerate(meses_idx, start=2):
                cell = ws.cell(row=linha, column=col)
                if idx == 0:
                    _ah_style(cell, valor="-", cor_fundo=COR_AH_G)
                else:
                    ah = _calcular_ah(valores[idx], valores[idx - 1])
                    if ah is None:
                        _ah_style(cell, valor="-", cor_fundo=COR_AH_G)
                    else:
                        cor_txt = COR_SUBIR if ah > 0 else COR_CAIR
                        _ah_style(cell, valor=ah, fmt='+0.0%;-0.0%;0.0%',
                                  cor_fundo=COR_AH_G, cor_texto=cor_txt)

            _ah_style(ws.cell(row=linha, column=col_total), valor="-", cor_fundo=COR_AH_G)
            _ah_style(ws.cell(row=linha, column=col_av), valor="-", cor_fundo=COR_AH_G)
            ws.row_dimensions[linha].height = 13
            linha += 1

        # ── TOTAL MÊS ─────────────────────────────────────────
        _total(ws.cell(row=linha, column=1), "TOTAL MÊS", align="left")
        total_ano = sum(totais_mes)
        for col, idx in enumerate(meses_idx, start=2):
            val = totais_mes[idx] if totais_mes[idx] else None
            _total(ws.cell(row=linha, column=col), val,
                   fmt='R$ #,##0.00' if val else None, bold=False)
        _total(ws.cell(row=linha, column=col_total), total_ano if total_ano else None,
               fmt='R$ #,##0.00', bold=True)
        _total(ws.cell(row=linha, column=col_av), 1.0, fmt='0.0%', bold=True)
        ws.row_dimensions[linha].height = 16
        linha += 1

        # ── A/H % geral ───────────────────────────────────────
        cell = ws.cell(row=linha, column=1, value="A/H %")
        cell.font      = Font(bold=True, color=COR_BRANCO, size=9)
        _fill(cell, COR_AH_G)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        for col, idx in enumerate(meses_idx, start=2):
            cell = ws.cell(row=linha, column=col)
            if idx == 0:
                _ah_style(cell, valor="-")
            else:
                ah = _calcular_ah(totais_mes[idx], totais_mes[idx - 1])
                if ah is None:
                    _ah_style(cell, valor="-")
                else:
                    cor_txt = COR_SUBIR if ah > 0 else COR_CAIR
                    _ah_style(cell, valor=ah, fmt='+0.0%;-0.0%;0.0%', cor_texto=cor_txt)

        _ah_style(ws.cell(row=linha, column=col_total), valor="-")
        _ah_style(ws.cell(row=linha, column=col_av), valor="-")
        ws.row_dimensions[linha].height = 15
        linha += 1

        return linha

    # ── Blocos ────────────────────────────────────────────────
    linha_atual = 1
    linha_atual = escrever_bloco(linha_atual, list(range(0, 6)),  MESES_ABREV[0:6])
    linha_atual += 1
    escrever_bloco(linha_atual, list(range(6, 12)), MESES_ABREV[6:12])

    # ── Largura das colunas ───────────────────────────────────
    ws.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 13
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 10

    # ── Configuração de página ────────────────────────────────
    ws.page_setup.orientation  = 'portrait'
    ws.page_setup.fitToPage    = True
    ws.page_setup.fitToWidth   = 1
    ws.page_setup.fitToHeight  = 0

    wb.save(caminho)